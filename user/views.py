from rest_framework import status
from django.contrib.auth import authenticate
from rest_framework.permissions import AllowAny
from .models import *
from .serializers import *
from django.dispatch import receiver
from django.db.models.signals import post_save
from rest_framework.decorators import permission_classes, api_view, throttle_classes
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from rest_framework.views import APIView
from rest_framework.response import Response
from .models import CityModel
import pandas as pd
import os
import logging
from django.db.models import Q, Count
from django.contrib.auth.hashers import make_password, check_password
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.views.decorators.cache import cache_page
from rest_framework.throttling import UserRateThrottle
from datetime import datetime
import io
import openpyxl
from collections import Counter
import jdatetime


class OncePerMinuteThrottle(UserRateThrottle):
    rate = "1/minute"


@api_view(["POST"])
@permission_classes([AllowAny])
def create_destination(request):
    s_data = request.data
    print(request.data)
    data_serializers = DestinationSerializers(data=s_data)
    if data_serializers.is_valid():
        data_serializers.save()
        return Response(data_serializers.data, status=status.HTTP_201_CREATED)
    return Response(data_serializers.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["GET"])
@permission_classes([AllowAny])
def get_all_destination(request):
    user_data = DestinationModel.objects.all()
    return Response(
        DestinationSerializers(user_data, many=True).data, status=status.HTTP_200_OK
    )


@api_view(["PATCH"])
@permission_classes([AllowAny])
@transaction.atomic
def edit_destination(request, id):
    person = get_object_or_404(DestinationModel, id=id)
    serializer = DestinationSerializers(person, data=request.data, partial=True)

    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["DELETE"])
def delete_anbar_request(request, id):
    try:
        destination_request = DestinationModel.objects.get(id=id)
    except DestinationModel.DoesNotExist:
        return Response({"error": "درخواست مورد نظر پیدا نشد."}, status=404)
    destination_request.delete()
    return Response({"message": "درخواست با موفقیت حذف شد."}, status=204)


@api_view(["POST"])
@permission_classes([AllowAny])
def create_person(request):
    s_data = request.data
    print(request.data)
    data_serializers = PersonSerializers(data=s_data)
    if data_serializers.is_valid():
        data_serializers.save()
        return Response(data_serializers.data, status=status.HTTP_201_CREATED)
    return Response(data_serializers.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["GET"])
@permission_classes([AllowAny])
@cache_page(60 * 15)  # کش برای 15 دقیقه
def get_all_person(request):
    user_data = (
        PersonelModel.objects.select_related(
            "unit", "work_type", "post_work", "sh_ostan", "burn_ostan"
        )
        .prefetch_related("child")
        .annotate(
            total_users=Count("id"),
            active_users=Count("id", filter=Q(is_online="A")),
            mission_users=Count("id", filter=Q(is_online="MA")),
            consultation_users=Count("id", filter=Q(is_online="MO")),
        )
        .first()
    )

    response_data = {
        "total_users": user_data.total_users if user_data else 0,
        "active_users": user_data.active_users if user_data else 0,
        "mission_users": user_data.mission_users if user_data else 0,
        "consultation_users": user_data.consultation_users if user_data else 0,
        "users": PersonGetSerializers(
            PersonelModel.objects.select_related(
                "unit", "work_type", "post_work", "sh_ostan", "burn_ostan"
            )
            .prefetch_related("child")
            .all(),
            many=True,
        ).data,
    }

    return Response(response_data, status=status.HTTP_200_OK)


@api_view(["GET"])
@permission_classes([AllowAny])
def get_postWork(request):
    user_data = PostWorkModel.objects.all()
    return Response(
        PostWorkSerializers(user_data, many=True).data, status=status.HTTP_200_OK
    )


@api_view(["GET"])
@permission_classes([AllowAny])
def get_workType(request):
    user_data = WorkTypeModel.objects.all()
    return Response(
        WorkTypeSerializers(user_data, many=True).data, status=status.HTTP_200_OK
    )


@api_view(["GET"])
@permission_classes([AllowAny])
def get_unit(request):
    user_data = UnitModel.objects.all()
    return Response(
        UnitSerializers(user_data, many=True).data, status=status.HTTP_200_OK
    )


@api_view(["GET"])
@permission_classes([AllowAny])
def get_ostan(request):
    user_data = OstanModel.objects.all()
    return Response(
        OstanSerializers(user_data, many=True).data, status=status.HTTP_200_OK
    )


@api_view(["GET"])
@permission_classes([AllowAny])
def get_shahrestan(request):
    user_data = CityModel.objects.all()
    return Response(
        CitySimpleSerializers(user_data, many=True).data, status=status.HTTP_200_OK
    )


@api_view(["GET"])
@permission_classes([AllowAny])
def get_shahrestan_by_ostanID(request, ostan_id):
    user_data = CityModel.objects.filter(ostan_id=ostan_id)
    return Response(
        CitySimpleSerializers(user_data, many=True).data, status=status.HTTP_200_OK
    )


@api_view(["GET"])
@permission_classes([AllowAny])
def get_child(request):
    user_data = ChildModel.objects.all()
    return Response(
        ChildSerializers(user_data, many=True).data, status=status.HTTP_200_OK
    )


@csrf_exempt
def delete_all_cities(request):
    if request.method == "DELETE":
        CityModel.objects.all().delete()
        return JsonResponse({"message": "تمام داده‌های شهرستان حذف شدند."}, status=200)
    return JsonResponse({"error": "متد غیرمجاز"}, status=405)


@api_view(["GET"])
@permission_classes([AllowAny])
@cache_page(60 * 15)  # کش برای 15 دقیقه
def get_flter_person(request):
    queryset = PersonelModel.objects.select_related(
        "unit", "work_type", "post_work", "sh_ostan", "burn_ostan"
    ).prefetch_related("child")

    filters = {}
    if request.GET.get("gender"):
        filters["gender"] = request.GET.get("gender")
    if request.GET.get("is_online"):
        filters["is_online"] = request.GET.get("is_online")
    if request.GET.get("shift"):
        filters["shift"] = request.GET.get("shift")
    if request.GET.get("person_code"):
        filters["person_code"] = request.GET.get("person_code")
    if request.GET.get("unit"):
        filters["unit_id"] = request.GET.get("unit")
    if request.GET.get("work_type"):
        filters["work_type_id"] = request.GET.get("work_type")
    if request.GET.get("post_work"):
        filters["post_work_id"] = request.GET.get("post_work")
    if request.GET.get("solder_select"):
        filters["solder_select"] = request.GET.get("solder_select")
    if request.GET.get("child_count"):
        filters["child_count"] = request.GET.get("child_count")

    try:
        if request.GET.get("date_employ"):
            filters["date_employ"] = request.GET.get("date_employ")
        if request.GET.get("date_employ__gte"):
            filters["date_employ__gte"] = request.GET.get("date_employ__gte")
        if request.GET.get("date_employ__lte"):
            filters["date_employ__lte"] = request.GET.get("date_employ__lte")
    except Exception as e:
        logging.error(f"خطا در فیلتر تاریخ: {e}")

    queryset = queryset.filter(**filters)
    serialized_data = PersonGetSerializers(queryset, many=True).data
    return Response(serialized_data, status=status.HTTP_200_OK)


class AddOstansFromExcel(APIView):
    def post(self, request, *args, **kwargs):
        file_path = "C:/Users/it/Desktop/Django/ostan.xlsx"

        try:
            df = pd.read_excel(file_path)
            df["ostan"] = df["ostan"].str.strip()

            new_ostans = []
            for ostan_name in df["ostan"].unique():
                if not OstanModel.objects.filter(name=ostan_name).exists():
                    new_ostans.append(OstanModel(name=ostan_name))

            if new_ostans:
                OstanModel.objects.bulk_create(new_ostans)

            return Response(
                {"message": f"{len(new_ostans)} استان با موفقیت اضافه شد."},
                status=status.HTTP_201_CREATED,
            )
        except Exception as e:
            logging.error(f"خطا در اضافه کردن استان‌ها: {e}")
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class AddCitiesFromExcel(APIView):
    def post(self, request, *args, **kwargs):
        file_path = "C:/Users/it/Desktop/Django/shahr.xlsx"

        try:
            df = pd.read_excel(file_path)
            df["City"] = df["City"].str.strip()

            new_cities = []
            for index, row in df.iterrows():
                state_id = row["State"]
                city_name = row["City"]

                try:
                    ostan = OstanModel.objects.get(id=state_id)
                    if not CityModel.objects.filter(
                        name=city_name, ostan=ostan
                    ).exists():
                        new_cities.append(CityModel(ostan=ostan, name=city_name))
                except OstanModel.DoesNotExist:
                    return Response(
                        {"error": f"استان با شناسه {state_id} وجود ندارد."},
                        status=status.HTTP_404_NOT_FOUND,
                    )

            if new_cities:
                CityModel.objects.bulk_create(new_cities)

            return Response(
                {"message": f"{len(new_cities)} شهر با موفقیت اضافه شد."},
                status=status.HTTP_201_CREATED,
            )
        except Exception as e:
            logging.error(f"خطا در اضافه کردن شهرها: {e}")
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ChangePasswordView(APIView):
    def post(self, request, *args, **kwargs):
        serializer = ChangePasswordSerializer(data=request.data)
        if serializer.is_valid():
            person_code = serializer.validated_data["person_code"]
            new_password = serializer.validated_data["new_password"]

            person = PersonelModel.objects.get(person_code=person_code)
            person.password = new_password
            person.save()

            return Response(
                {"status": True, "message": "رمز عبور با موفقیت تغییر کرد."},
                status=status.HTTP_200_OK,
            )

        # دریافت اولین خطا و تعیین status
        errors = serializer.errors
        first_error_field = list(errors.keys())[0]
        first_error = errors[first_error_field]

        # اگه پیام به صورت dict بود، status رو ازش استخراج می‌کنیم
        custom_status = 400
        if isinstance(first_error, dict):
            message = first_error.get("message", "خطای نامشخصص")
            custom_status = first_error.get("status", 400)
        else:
            message = first_error[0]

        return Response(
            {"status": False, "field": first_error_field, "message": message},
            status=custom_status,
        )


@api_view(["PATCH"])
@permission_classes([AllowAny])
@transaction.atomic
def edit_user(request, id):
    person = get_object_or_404(PersonelModel, person_code=id)
    serializer = PersonSerializers(person, data=request.data, partial=True)

    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["PATCH"])
@permission_classes([AllowAny])
@transaction.atomic
def edit_unit(request, id):
    unit = get_object_or_404(UnitModel, id=id)
    serializer = UnitSerializers(unit, data=request.data, partial=True)

    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["PATCH"])
@permission_classes([AllowAny])
@transaction.atomic
def edit_postwork(request, id):
    postwork = get_object_or_404(PostWorkModel, id=id)
    serializer = PostWorkSerializers(postwork, data=request.data, partial=True)

    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["PATCH"])
@permission_classes([AllowAny])
@transaction.atomic
def edit_worktype(request, id):
    worktype = get_object_or_404(WorkTypeModel, id=id)
    serializer = WorkTypeSerializers(worktype, data=request.data, partial=True)

    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class FivePerMinuteThrottle(UserRateThrottle):
    rate = "10/min"


@api_view(["POST"])
@permission_classes([AllowAny])
@throttle_classes([FivePerMinuteThrottle])
def login_view(request):
    serializer = LoginSerializer(data=request.data)
    if serializer.is_valid():
        person_code = serializer.validated_data["person_code"]
        password = serializer.validated_data["password"]
        try:
            user = PersonelModel.objects.get(person_code=person_code)
            if password == user.password:
                user_data = PersonGetSerializers(user).data
                return Response(user_data, status=status.HTTP_200_OK)
            else:
                return Response(
                    {"error": "رمز عبور اشتباه است"},
                    status=status.HTTP_401_UNAUTHORIZED,
                )
        except PersonelModel.DoesNotExist:
            return Response(
                {"error": "کاربری با این شماره پرسنلی یافت نشد"},
                status=status.HTTP_404_NOT_FOUND,
            )
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["POST"])
@permission_classes([AllowAny])
def signup_view(request):
    serializer = SignUpSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "ثبت‌ نام با موفقیت انجام شد"}, status=status.HTTP_201_CREATED
        )
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class GetUserDataView(APIView):
    def post(self, request):
        person_code = request.data.get("person_code")
        if not person_code:
            return Response(
                {"detail": "کد پرسنلی الزامی است."}, status=status.HTTP_400_BAD_REQUEST
            )
        try:
            person = PersonelModel.objects.get(person_code=person_code)
        except PersonelModel.DoesNotExist:
            return Response(
                {"detail": "کاربری با این کد پرسنلی وجود ندارد."},
                status=status.HTTP_404_NOT_FOUND,
            )

        serializer = PersonGetSerializers(person)
        return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(["GET"])
@permission_classes([AllowAny])
def download_users_excel(request):
    try:
        # دریافت تمام کاربران با اطلاعات مرتبط
        users = PersonelModel.objects.select_related(
            "unit", "work_type", "post_work", "sh_ostan", "burn_ostan"
        ).prefetch_related("child")

        # ایجاد فایل اکسل در حافظه
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users"

        # تعریف هدرها
        headers = [
            "Personnel Code",
            "First Name",
            "Last Name",
            "Gender",
            "Status",
            "Shift",
            "Unit",
            "Work Type",
            "Job Position",
            "Service Province",
            "Birth Province",
            "Children Count",
            "Male Children",
            "Female Children",
            "Employment Date",
            "Military Status",
            "Exemption Reason",
            "Education",
            "Education Field",
            "University",
            "Institute",
            "Document in File",
            "Marital Status",
            "Spouse Name",
            "Spouse Birth Date",
            "Insurance Number",
            "Insurance Days",
            "Insurance Years",
            "Bank Account",
            "IBAN",
            "Address",
            "Postal Code",
            "Settlement Date",
            "Settlement Reason",
            "In Settlement",
            "Issue Date",
            "Marriage Date",
            "Access Level",
            "Phone Number",
            "National Code",
            "ID Number",
            "ID Serial",
            "Father Name",
            "Birth Date",
            "Is Active",
            "Create Date",
            "Update Date",
        ]

        # نوشتن هدرها
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)

        # Define mappings for choices based on models.py
        GENDER_MAP = dict(GENDER)
        ISONLINE_MAP = dict(ISONLINE)
        SHIFT_WORK_MAP = dict(SHIFT_WORK)
        SOLDER_SELECT_MAP = dict(SOLDER_SELECT)
        EDUCATION_SELECT_MAP = dict(EDUCATION_SELECT)
        ACCESS_SELECT_MAP = dict(ACCESS_SELECT)

        # نوشتن داده‌ها
        for row_num, user in enumerate(users, 2):
            # Helper function to safely format dates/datetimes
            def format_datetime(dt):
                if dt:
                    # Check if it's jdatetime.date or jdatetime.datetime
                    if hasattr(
                        dt, "strftime"
                    ):  # Standard datetime/date or jdatetime.datetime
                        try:
                            # Attempt standard formatting first
                            return (
                                dt.strftime("%Y-%m-%d %H:%M:%S")
                                if hasattr(dt, "hour")
                                else dt.strftime("%Y-%m-%d")
                            )
                        except (
                            ValueError
                        ):  # Handle potential jdatetime specific formatting if needed
                            # Assuming jdatetime objects might need specific formatting if strftime fails
                            # Or simply convert to string
                            return str(dt)
                    else:  # Fallback for other types or if strftime is not available
                        return str(dt)
                return ""

            data = [
                user.person_code,
                user.first_name,
                user.last_name,
                GENDER_MAP.get(user.gender, user.gender),  # Use map for Gender
                ISONLINE_MAP.get(user.is_online, user.is_online),  # Use map for Status
                SHIFT_WORK_MAP.get(user.shift, user.shift),  # Use map for Shift
                user.unit.name if user.unit else "",
                user.work_type.name if user.work_type else "",
                user.post_work.name if user.post_work else "",
                user.sh_ostan.name if user.sh_ostan else "",
                user.burn_ostan.name if user.burn_ostan else "",
                user.child_count,
                user.boy_child_count,
                user.girl_child_count,
                format_datetime(user.date_employ),  # Format date
                SOLDER_SELECT_MAP.get(
                    user.solder_select, user.solder_select
                ),  # Use map for Military Status
                user.reason_ex,
                EDUCATION_SELECT_MAP.get(
                    user.education, user.education
                ),  # Use map for Education
                user.education_field,
                user.univercity,
                user.Institute,
                "Yes" if user.ev_file else "No",
                "Married" if user.is_marid else "Single",
                user.wife_name,
                format_datetime(user.wife_birthday),  # Format date
                user.insurance_number,
                user.insurance_history_day,
                user.insurance_history_year,
                user.bank_number,
                user.bank_shaba,
                user.address,
                user.post_code,
                format_datetime(user.settlement_date),  # Format date
                user.settlement_description,
                "Yes" if user.settlement_steps else "No",
                format_datetime(user.issue_date),  # Format date
                format_datetime(user.married_date),  # Format date
                ACCESS_SELECT_MAP.get(
                    user.access, user.access
                ),  # Use map for Access Level
                user.phone_number,
                user.melli_code,
                user.sh_code,
                user.sh_s_code,
                user.father_name,
                format_datetime(user.birthday),  # Format date
                "Yes" if user.is_active else "No",
                format_datetime(user.create_at),  # Format datetime
                format_datetime(user.update_at),  # Format datetime
            ]

            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                # Ensure value is not None before assigning
                cell.value = value if value is not None else ""

        # تنظیم عرض ستون‌ها
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        # ذخیره فایل
        wb.save(output)
        output.seek(0)

        # تنظیم هدرهای پاسخ
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename=users_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

        return response

    except Exception as e:
        # Log the specific error for better debugging
        logging.error(f"خطا در ایجاد فایل اکسل: {e}", exc_info=True)
        return Response(
            {"error": f"خطا در ایجاد فایل اکسل: {e}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([AllowAny])
def get_user_active_status(request, person_code):
    try:
        # Attempt to find the user by person_code
        user = PersonelModel.objects.get(person_code=person_code)
        # Return the is_active status
        return Response({"is_active": user.is_active}, status=status.HTTP_200_OK)
    except PersonelModel.DoesNotExist:
        # Handle case where user is not found
        return Response(
            {"error": "کاربری با این شماره پرسنلی یافت نشد"},
            status=status.HTTP_404_NOT_FOUND,
        )
    except Exception as e:
        # Handle other potential errors
        logging.error(
            f"Error checking user active status for {person_code}: {e}", exc_info=True
        )
        return Response(
            {"error": "خطای داخلی سرور"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
